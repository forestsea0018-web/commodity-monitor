#!/usr/bin/env python3
"""
半导体全产业链标的 Excel 生成器 — 三张工作表
  Sheet1: 半导体制造设备  (按工序阶段 + 公司×工艺参与矩阵补全)
  Sheet2: 半导体关键材料  (12大卡脖子材料)
  Sheet3: AI算力产业链    (2026年核心赛道20大细分)
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# 样式辅助
# ═══════════════════════════════════════════════════════════════════════════════
def hf(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def tf(bold=False, color="000000", size=9, name="微软雅黑", italic=False):
    return Font(bold=bold, color=color, size=size, name=name, italic=italic)

thin  = Side(style="thin",   color="BFBFBF")
med   = Side(style="medium", color="808080")
BORDER      = Border(left=thin, right=thin, top=thin,  bottom=thin)
BORDER_TOP  = Border(left=thin, right=thin, top=med,   bottom=thin)

C = Alignment(horizontal="center", vertical="center", wrap_text=True)
L = Alignment(horizontal="left",   vertical="center", wrap_text=True)

HEADER_FILL = hf("1F3864")
HEADER_FONT = tf(bold=True, color="FFFFFF", size=10)

# 价格列标题
PRICE_HEADERS = [
    "昨日收盘价", "实时现价", "2025年末价格", "Q1最终价格",
    "过去三个月均价", "过去半年趋势", "MS目标价", "高盛目标价", "摩根大通",
]
PRICE_COLS = len(PRICE_HEADERS)

# 公共前缀列宽（不同sheet会加自己的固定列）
def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def write_header_row(ws, row, headers, row_height=34):
    ws.row_dimensions[row].height = row_height
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, C, BORDER

def write_title(ws, text, ncols, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws["A1"]
    c.value, c.font, c.fill, c.alignment = (
        text, tf(bold=True, color="FFFFFF", size=14),
        hf("1F3864"), C
    )
    ws.row_dimensions[1].height = 32
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c2 = ws["A2"]
        c2.value, c2.font, c2.fill, c2.alignment = (
            subtitle, tf(italic=True, color="595959", size=9),
            hf("D6E4F7"), L
        )
        ws.row_dimensions[2].height = 18
        return 3   # header row
    return 2


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — 半导体制造设备
# ═══════════════════════════════════════════════════════════════════════════════
# (阶段, 步骤号, 步骤名, 类型, 说明, 公司名, 代码, 市场, 细分行业, 币种, 参与度)
# 参与度: ★主导/垄断(>50%)  ●核心参与  ○参与
EQUIP_DATA = [
    # ── 0. EDA设计 ─────────────────────────────────────────────────────────────
    ("0. Design EDA设计","0","EDA软件","SEQ","逻辑综合→P&R→时序签核→GDSII",
     "Cadence Design Systems","CDNS","NASDAQ","EDA软件","USD","★"),
    ("0. Design EDA设计","0","EDA软件","SEQ","逻辑综合→P&R→时序签核→GDSII",
     "Synopsys","SNPS","NASDAQ","EDA软件","USD","★"),

    # ── 0.5 掩膜制备 ───────────────────────────────────────────────────────────
    ("0.5 Mask 掩膜制备","0.5","掩膜写入","SEQ","E-beam写入掩膜 (非上市，东芝子公司)",
     "NuFlare Technology","非上市","—","掩膜写入设备","JPY","●"),
    ("0.5 Mask 掩膜制备","0.5","掩膜写入","SEQ","E-beam写入设备",
     "JEOL","6951.T","东京证交所","电子束设备/检测","JPY","○"),
    ("0.5 Mask 掩膜制备","0.5","掩膜检测","SEQ","掩膜缺陷检测，EUV掩膜绝对垄断",
     "Lasertec","6920.T","东京证交所","掩膜检测设备","JPY","★"),
    ("0.5 Mask 掩膜制备","0.5","空白掩膜板","SEQ","空白板材料供应，全球垄断",
     "Hoya Corporation","7741.T","东京证交所","空白掩膜板/光学元件","JPY","★"),
    ("0.5 Mask 掩膜制备","0.5","空白掩膜板","SEQ","空白板材料供应",
     "Shin-Etsu Chemical 信越化学","4063.T","东京证交所","空白掩膜板/硅片","JPY","●"),

    # ── 1. FEOL 前端晶体管 ×20-30层 ──────────────────────────────────────────
    # ①Coat
    ("1. FEOL 前端×20-30层","①","Coat 涂布","LOOP×N","光阻涂布 Resist Coat，TEL 93%垄断",
     "Tokyo Electron / TEL 东京电子","8035.T","东京证交所","涂布/显影/热处理设备","JPY","★"),
    ("1. FEOL 前端×20-30层","①","Coat 涂布","LOOP×N","光阻涂布，韩国三星设备子公司(非上市)",
     "SEMES","非上市","—","涂布/清洗设备","KRW","○"),

    # ②Expose
    ("1. FEOL 前端×20-30层","②","Expose 曝光","LOOP×N","EUV全球垄断，DUV主力",
     "ASML Holding","ASML","NASDAQ","光刻机 EUV/DUV","USD","★"),
    ("1. FEOL 前端×20-30层","②","Expose 曝光","LOOP×N","i-line / DUV准分子激光光刻",
     "Canon 佳能","7751.T","东京证交所","光刻机 i-line/DUV","JPY","●"),
    ("1. FEOL 前端×20-30层","②","Expose 曝光","LOOP×N","DUV准分子激光光刻",
     "Nikon 尼康","7731.T","东京证交所","光刻机 DUV","JPY","○"),

    # ③Develop
    ("1. FEOL 前端×20-30层","③","Develop 显影","LOOP×N","曝光后显影 Post-Expose Develop",
     "Tokyo Electron / TEL 东京电子","8035.T","东京证交所","涂布/显影设备","JPY","★"),

    # ④a Etch
    ("1. FEOL 前端×20-30层","④a","Etch 蚀刻","PAR","干法蚀刻，存储市场~70%份额",
     "Lam Research","LRCX","NASDAQ","蚀刻设备","USD","★"),
    ("1. FEOL 前端×20-30层","④a","Etch 蚀刻","PAR","干法蚀刻，逻辑市场~30%份额",
     "Tokyo Electron / TEL 东京电子","8035.T","东京证交所","蚀刻/CVD设备","JPY","●"),
    ("1. FEOL 前端×20-30层","④a","Etch 蚀刻","PAR","中国本土替代，科创板",
     "中微公司 AMEC","688012.SS","上交所科创板","蚀刻设备","CNY","○"),
    ("1. FEOL 前端×20-30层","④a","Etch 蚀刻","PAR","中国本土替代",
     "北方华创 Naura","002371.SZ","深交所","蚀刻/CVD/PVD设备","CNY","○"),

    # ④b Ion Implant
    ("1. FEOL 前端×20-30层","④b","Ion Impl. 离子注入","PAR","离子注入掺杂，AMAT全球主导",
     "Applied Materials / AMAT 应用材料","AMAT","NASDAQ","离子注入/CVD/PVD/CMP设备","USD","★"),

    # ⑤Thermal
    ("1. FEOL 前端×20-30层","⑤","Thermal 热处理","LOOP×N","热处理/退火 Furnace/RTP，TEL 53%",
     "Tokyo Electron / TEL 东京电子","8035.T","东京证交所","热处理炉设备","JPY","★"),
    ("1. FEOL 前端×20-30层","⑤","Thermal 热处理","LOOP×N","热处理/退火，KOKUSAI 30%",
     "Kokusai Electric","6525.T","东京证交所","热处理炉设备","JPY","●"),
    ("1. FEOL 前端×20-30层","⑤","Thermal 热处理","LOOP×N","ALD/CVD及热处理",
     "ASM International / ASMI","ASM.AS","阿姆斯特丹","ALD/热处理设备","EUR","○"),
    ("1. FEOL 前端×20-30层","⑤","Thermal 热处理","LOOP×N","中国本土替代，热处理/CVD",
     "北方华创 Naura","002371.SZ","深交所","热处理/CVD设备","CNY","○"),

    # ⑥Clean
    ("1. FEOL 前端×20-30层","⑥","Clean 清洗","LOOP×N","单片清洗>50%，占全工艺步骤>30%",
     "SCREEN Holdings 迪恩士","7735.T","东京证交所","清洗设备","JPY","★"),
    ("1. FEOL 前端×20-30层","⑥","Clean 清洗","LOOP×N","单片清洗，中国本土替代",
     "ACM Research / ACMR 盛美","ACMR","NASDAQ","清洗/电镀设备","USD","○"),
    ("1. FEOL 前端×20-30层","⑥","Clean 清洗","LOOP×N","清洗设备，三星设备子公司(非上市)",
     "SEMES","非上市","—","清洗设备","KRW","○"),
    ("1. FEOL 前端×20-30层","⑥","Clean 清洗","LOOP×N","清洗设备",
     "Tokyo Electron / TEL 东京电子","8035.T","东京证交所","清洗设备","JPY","●"),

    # ⑦Inspect
    ("1. FEOL 前端×20-30层","⑦","Inspect 检测","LOOP×N","在线缺陷检测，KLA >50%",
     "KLA Corporation","KLAC","NASDAQ","量检测/计量设备","USD","★"),
    ("1. FEOL 前端×20-30层","⑦","Inspect 检测","LOOP×N","电子束检测/量测",
     "Hitachi High-Tech 日立高新","6501.T*","东京证交所(日立母公司)","电子束检测设备","JPY","●"),
    ("1. FEOL 前端×20-30层","⑦","Inspect 检测","LOOP×N","电子显微镜/材料分析",
     "Thermo Fisher Scientific","TMO","NYSE","电子显微镜/分析仪器","USD","○"),
    ("1. FEOL 前端×20-30层","⑦","Inspect 检测","LOOP×N","电子显微镜/束流设备",
     "JEOL","6951.T","东京证交所","电子束设备/检测","JPY","○"),

    # ── 2. BEOL 前端互连线 ×7-15铜层 ─────────────────────────────────────────
    # ⑧CVD
    ("2. BEOL 互连×7-15铜层","⑧","CVD/ALD 成膜","LOOP×N","介电膜/阻挡层 CVD/ALD Deposition",
     "Applied Materials / AMAT 应用材料","AMAT","NASDAQ","CVD/ALD设备","USD","●"),
    ("2. BEOL 互连×7-15铜层","⑧","CVD/ALD 成膜","LOOP×N","介电膜/ALD，TEL强势品类",
     "Tokyo Electron / TEL 东京电子","8035.T","东京证交所","CVD/ALD设备","JPY","●"),
    ("2. BEOL 互连×7-15铜层","⑧","CVD/ALD 成膜","LOOP×N","CVD/ALD，Lam参与",
     "Lam Research","LRCX","NASDAQ","CVD/ALD设备","USD","●"),
    ("2. BEOL 互连×7-15铜层","⑧","CVD/ALD 成膜","LOOP×N","批次CVD炉管，KOKUSAI强势",
     "Kokusai Electric","6525.T","东京证交所","批次CVD炉管设备","JPY","●"),
    ("2. BEOL 互连×7-15铜层","⑧","CVD/ALD 成膜","LOOP×N","ALD设备，ASMI全球领先",
     "ASM International / ASMI","ASM.AS","阿姆斯特丹","ALD设备","EUR","●"),

    # ⑨PVD
    ("2. BEOL 互连×7-15铜层","⑨","PVD 溅射","LOOP×N","金属溅射 Metal Sputter，AMAT主导",
     "Applied Materials / AMAT 应用材料","AMAT","NASDAQ","PVD溅射设备","USD","★"),
    ("2. BEOL 互连×7-15铜层","⑨","PVD 溅射","LOOP×N","PVD溅射，中国本土替代",
     "北方华创 Naura","002371.SZ","深交所","PVD溅射设备","CNY","○"),
    ("2. BEOL 互连×7-15铜层","⑨","PVD 溅射","LOOP×N","PVD溅射",
     "Ulvac","6728.T","东京证交所","PVD溅射/真空设备","JPY","●"),

    # ⑪Plate
    ("2. BEOL 互连×7-15铜层","⑪","Plate 铜电镀","LOOP×N","Cu Electroplating，Lam ~80%",
     "Lam Research","LRCX","NASDAQ","铜电镀设备","USD","★"),
    ("2. BEOL 互连×7-15铜层","⑪","Plate 铜电镀","LOOP×N","铜电镀，AMAT参与",
     "Applied Materials / AMAT 应用材料","AMAT","NASDAQ","铜电镀/ECD设备","USD","●"),
    ("2. BEOL 互连×7-15铜层","⑪","Plate 铜电镀","LOOP×N","铜电镀，中国本土替代",
     "ACM Research / ACMR 盛美","ACMR","NASDAQ","铜电镀/清洗设备","USD","○"),
    ("2. BEOL 互连×7-15铜层","⑪","Plate 铜电镀","LOOP×N","铜电镀，Ebara参与",
     "荏原制作所 Ebara","6361.T","东京证交所","铜电镀/CMP设备","JPY","●"),

    # ⑫CMP
    ("2. BEOL 互连×7-15铜层","⑫","CMP 化学机械抛光","LOOP×N","Cu CMP，AMAT >50%",
     "Applied Materials / AMAT 应用材料","AMAT","NASDAQ","CMP设备","USD","★"),
    ("2. BEOL 互连×7-15铜层","⑫","CMP 化学机械抛光","LOOP×N","CMP，Ebara全球第2",
     "荏原制作所 Ebara","6361.T","东京证交所","CMP/电镀设备","JPY","●"),

    # ⑬Clean+Insp
    ("2. BEOL 互连×7-15铜层","⑬","Clean+Insp","LOOP×N","同FEOL⑥⑦",
     "SCREEN Holdings 迪恩士","7735.T","东京证交所","清洗设备","JPY","★"),
    ("2. BEOL 互连×7-15铜层","⑬","Clean+Insp","LOOP×N","同FEOL⑦",
     "KLA Corporation","KLAC","NASDAQ","量检测设备","USD","★"),

    # ── 3. Mid-end 中端先进封装前道 ───────────────────────────────────────────
    # ⑭Grind
    ("3. Mid-end 先进封装前道","⑭","Grind 减薄","SEQ","晶圆背面减薄 <30μm，Disco #1",
     "Disco Corporation 迪斯科","6146.T","东京证交所","晶圆减薄/切割设备","JPY","★"),
    ("3. Mid-end 先进封装前道","⑭","Grind 减薄","SEQ","晶圆减薄，Tokyo Seimitsu #2",
     "Tokyo Seimitsu / Accretech 东京精密","7729.T","东京证交所","晶圆减薄/量测/探针","JPY","●"),
    ("3. Mid-end 先进封装前道","⑭","Grind 减薄","SEQ","减薄后检测",
     "KLA Corporation","KLAC","NASDAQ","量检测设备","USD","○"),

    # ⑮a Hybrid Bond
    ("3. Mid-end 先进封装前道","⑮a","Hybrid Bond 混合键合","NEW",
     "W2W/C2W混合键合(CoWoS/SoIC)探索中，TEL参与",
     "Tokyo Electron / TEL 东京电子","8035.T","东京证交所","混合键合设备","JPY","●"),
    ("3. Mid-end 先进封装前道","⑮a","Hybrid Bond 混合键合","NEW",
     "W2W键合设备，EVG私有奥地利企业",
     "EV Group / EVG","非上市","—","W2W键合设备","EUR","●"),

    # ⑮b TSV
    ("3. Mid-end 先进封装前道","⑮b","TSV 硅通孔","SEQ","TSV Deep Etch，Lam主导",
     "Lam Research","LRCX","NASDAQ","TSV蚀刻/CVD设备","USD","●"),
    ("3. Mid-end 先进封装前道","⑮b","TSV 硅通孔","SEQ","TSV铜填充电镀",
     "荏原制作所 Ebara","6361.T","东京证交所","TSV电镀设备","JPY","●"),
    ("3. Mid-end 先进封装前道","⑮b","TSV 硅通孔","SEQ","TSV CMP",
     "Applied Materials / AMAT 应用材料","AMAT","NASDAQ","TSV CMP设备","USD","●"),

    # ⑯RDL
    ("3. Mid-end 先进封装前道","⑯","RDL 再布线","SEQ","重分布层 Redistribution Layer",
     "Tokyo Electron / TEL 东京电子","8035.T","东京证交所","RDL涂布/曝光设备","JPY","●"),
    ("3. Mid-end 先进封装前道","⑯","RDL 再布线","SEQ","RDL清洗",
     "SCREEN Holdings 迪恩士","7735.T","东京证交所","RDL清洗设备","JPY","●"),
    ("3. Mid-end 先进封装前道","⑯","RDL 再布线","SEQ","RDL铜电镀",
     "荏原制作所 Ebara","6361.T","东京证交所","RDL电镀设备","JPY","●"),
    ("3. Mid-end 先进封装前道","⑯","RDL 再布线","SEQ","RDL铜电镀，中国本土替代",
     "ACM Research / ACMR 盛美","ACMR","NASDAQ","RDL电镀/清洗设备","USD","○"),

    # ⑰Probe Test
    ("3. Mid-end 先进封装前道","⑰","Probe Test 探针测试","SEQ","晶圆级探针测试 Wafer Probe",
     "Tokyo Seimitsu / Accretech 东京精密","7729.T","东京证交所","探针台","JPY","●"),
    ("3. Mid-end 先进封装前道","⑰","Probe Test 探针测试","SEQ","探针卡制造商 #1",
     "Micronics Japan","6871.T","东京证交所","探针卡","JPY","●"),
    ("3. Mid-end 先进封装前道","⑰","Probe Test 探针测试","SEQ","探针卡制造商 #2",
     "FormFactor","FORM","NASDAQ","探针卡","USD","●"),

    # ── 4. Back-end 后端封装测试 ──────────────────────────────────────────────
    # ⑱Dice
    ("4. Back-end 后端封装测试","⑱","Dice 切割","SEQ","晶圆切割 Wafer Dicing，Disco >70%",
     "Disco Corporation 迪斯科","6146.T","东京证交所","晶圆切割设备","JPY","★"),
    ("4. Back-end 后端封装测试","⑱","Dice 切割","SEQ","晶圆切割，Tokyo Seimitsu #2",
     "Tokyo Seimitsu / Accretech 东京精密","7729.T","东京证交所","晶圆切割/量测","JPY","○"),
    ("4. Back-end 后端封装测试","⑱","Dice 切割","SEQ","切割后检测",
     "KLA Corporation","KLAC","NASDAQ","切割后检测","USD","○"),

    # ⑲a FC Bond
    ("4. Back-end 后端封装测试","⑲a","FC Bond 倒装键合","PAR",
     "倒装键合/先进封装设备，ASMPT全球领先",
     "ASMPT","0522.HK","港交所","倒装键合/贴片设备","HKD","●"),
    ("4. Back-end 后端封装测试","⑲a","FC Bond 倒装键合","PAR",
     "TCB热压键合，Besi全球 #1",
     "BE Semiconductor / Besi","BESI.AS","阿姆斯特丹","倒装键合设备","EUR","★"),
    ("4. Back-end 后端封装测试","⑲a","FC Bond 倒装键合","PAR",
     "TC键合(HBM关键)，Hanmi核心供应商",
     "Hanmi Semiconductor 韩美半导体","042700.KQ","韩国KOSDAQ","TC键合/视觉检测","KRW","●"),
    ("4. Back-end 后端封装测试","⑲a","FC Bond 倒装键合","PAR","键合/封装设备",
     "Shibaura Mechatronics 芝浦机械","6590.T","东京证交所","键合/封装设备","JPY","●"),

    # ⑲b Wire Bond
    ("4. Back-end 后端封装测试","⑲b","Wire Bond 引线键合","PAR","金/铜引线焊接，K&S全球垄断",
     "Kulicke & Soffa / K&S","KLIC","NASDAQ","引线键合设备","USD","★"),
    ("4. Back-end 后端封装测试","⑲b","Wire Bond 引线键合","PAR","引线焊接，日本第2",
     "Shinkawa 新川","6274.T","东京证交所","引线键合设备","JPY","●"),

    # ⑳Mold
    ("4. Back-end 后端封装测试","⑳","Mold 模塑","SEQ","压缩模塑封装，TOWA全球垄断",
     "TOWA Corporation","6315.T","东京证交所","压缩模塑设备","JPY","★"),
    ("4. Back-end 后端封装测试","⑳","Mold 模塑","SEQ","压缩模塑，日本第2",
     "Apic Yamada","6300.T","东京证交所","压缩模塑/引线框架设备","JPY","●"),

    # ㉑a ATE Test
    ("4. Back-end 后端封装测试","㉑a","ATE Test 芯片测试","SEQ","存储/SoC ATE测试，Advantest #1",
     "Advantest","6857.T","东京证交所","ATE测试设备","JPY","★"),
    ("4. Back-end 后端封装测试","㉑a","ATE Test 芯片测试","SEQ","ATE测试，Teradyne #2",
     "Teradyne 泰瑞达","TER","NASDAQ","ATE测试设备","USD","●"),

    # ㉑b SLT
    ("4. Back-end 后端封装测试","㉑b","SLT 系统测试","NEW","系统级测试 新兴赛道",
     "Advantest","6857.T","东京证交所","ATE/SLT测试设备","JPY","●"),
    ("4. Back-end 后端封装测试","㉑b","SLT 系统测试","NEW","SLT测试",
     "Teradyne 泰瑞达","TER","NASDAQ","ATE/SLT测试设备","USD","●"),
    ("4. Back-end 后端封装测试","㉑b","SLT 系统测试","NEW","测试测量，Keysight参与",
     "Keysight Technologies","KEYS","NYSE","测试测量仪器","USD","●"),

    # ㉒Burn-in
    ("4. Back-end 后端封装测试","㉒","Burn-in 老化筛选","SEQ","可靠性老化筛选",
     "Advantest","6857.T","东京证交所","老化/测试设备","JPY","●"),
    ("4. Back-end 后端封装测试","㉒","Burn-in 老化筛选","SEQ","可靠性老化筛选",
     "Teradyne 泰瑞达","TER","NASDAQ","老化/测试设备","USD","●"),
]

STAGE_COLORS = {
    "0. Design EDA设计":       "4472C4",
    "0.5 Mask 掩膜制备":       "7030A0",
    "1. FEOL 前端×20-30层":    "C00000",
    "2. BEOL 互连×7-15铜层":   "833C00",
    "3. Mid-end 先进封装前道":  "ED7D31",
    "4. Back-end 后端封装测试": "375623",
}
TYPE_COLORS = {"SEQ":"00B0F0","LOOP×N":"92D050","PAR":"FF0000","NEW":"FFC000"}
PARTI_COLORS = {"★":"FF0000","●":"4472C4","○":"BFBFBF"}


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — 半导体关键材料 (12大卡脖子)
# ═══════════════════════════════════════════════════════════════════════════════
# (编号, 材料类别, 重要性说明, 公司名, 代码, 市场, 细分行业, 币种)
MATERIAL_DATA = [
    # ①磷化镓/化合物半导体衬底
    ("①","磷化镓/化合物半导体衬底","全球供应极度紧张，高速光芯片/毫米波雷达核心底材",
     "云南锗业","002428.SZ","深交所","锗/化合物半导体材料","CNY"),
    ("①","磷化镓/化合物半导体衬底","全球供应极度紧张，高速光芯片/毫米波雷达核心底材",
     "有研新材","600206.SH","上交所","半导体新材料/靶材","CNY"),
    ("①","磷化镓/化合物半导体衬底","全球供应极度紧张，高速光芯片/毫米波雷达核心底材",
     "光智科技","300161.SZ","深交所创业板","化合物半导体光学材料","CNY"),

    # ②光刻胶
    ("②","光刻胶 Photoresist","JSR/信越垄断，国内自给率<10%，高端光刻胶极度短缺",
     "彤程新材","603650.SH","上交所","光刻胶/半导体材料","CNY"),
    ("②","光刻胶 Photoresist","JSR/信越垄断，国内自给率<10%，高端光刻胶极度短缺",
     "南大光电","300346.SZ","深交所创业板","ArF光刻胶/前驱体","CNY"),
    ("②","光刻胶 Photoresist","JSR/信越垄断，国内自给率<10%，高端光刻胶极度短缺",
     "上海新阳","300236.SZ","深交所创业板","光刻胶/电镀液","CNY"),

    # ③碳化硅
    ("③","碳化硅 SiC","800V高压快充/新能源车驱动核心材料，今年价格涨幅>50%",
     "天岳先进","688234.SH","上交所科创板","SiC衬底","CNY"),
    ("③","碳化硅 SiC","800V高压快充/新能源车驱动核心材料，今年价格涨幅>50%",
     "露笑科技","002617.SZ","深交所","SiC外延/器件","CNY"),
    ("③","碳化硅 SiC","800V高压快充/新能源车驱动核心材料，今年价格涨幅>50%",
     "三安光电","600703.SH","上交所","SiC/化合物半导体器件","CNY"),

    # ④ABF载板
    ("④","ABF载板 Substrate","CPU/GPU封装基板，ABF膜日本垄断，价格涨幅>70%",
     "深南电路","002916.SZ","深交所","PCB/封装基板","CNY"),
    ("④","ABF载板 Substrate","CPU/GPU封装基板，ABF膜日本垄断，价格涨幅>70%",
     "兴森科技","002436.SZ","深交所","HDI/封装基板","CNY"),
    ("④","ABF载板 Substrate","CPU/GPU封装基板，ABF膜日本垄断，价格涨幅>70%",
     "崇达技术","002815.SZ","深交所","PCB/高密度板","CNY"),

    # ⑤钽电容
    ("⑤","钽电容 Tantalum Capacitor","积小/容大/耐高温，军工/航天/AI服务器刚需，涨幅>80%",
     "宏达电子","300726.SZ","深交所创业板","钽电容/被动元件","CNY"),
    ("⑤","钽电容 Tantalum Capacitor","积小/容大/耐高温，军工/航天/AI服务器刚需，涨幅>80%",
     "火炬电子","600650.SH","上交所","钽电容/军用元件","CNY"),
    ("⑤","钽电容 Tantalum Capacitor","积小/容大/耐高温，军工/航天/AI服务器刚需，涨幅>80%",
     "振华科技","000733.SZ","深交所","军用电子元件","CNY"),

    # ⑥高端PCB载板
    ("⑥","高端PCB载板","AI服务器需求爆发，今年大涨60%，算力硬件关键基材",
     "生益科技","600183.SH","上交所","覆铜板/PCB材料","CNY"),
    ("⑥","高端PCB载板","AI服务器需求爆发，今年大涨60%，算力硬件关键基材",
     "华正新材","603186.SH","上交所","高频PCB材料/覆铜板","CNY"),
    ("⑥","高端PCB载板","AI服务器需求爆发，今年大涨60%，算力硬件关键基材",
     "沪电股份","002463.SZ","深交所","高端PCB/通信板","CNY"),

    # ⑦电子级硫酸
    ("⑦","电子级硫酸 Electronic Grade H₂SO₄","芯片制造湿法清洗/刻蚀核心化学品，Stella垄断，涨幅>50%",
     "江化微","603078.SH","上交所","电子级硫酸/湿化学品","CNY"),
    ("⑦","电子级硫酸 Electronic Grade H₂SO₄","芯片制造湿法清洗/刻蚀核心化学品，Stella垄断，涨幅>50%",
     "多氟多","002407.SZ","深交所","电子级氟化物/锂电材料","CNY"),
    ("⑦","电子级硫酸 Electronic Grade H₂SO₄","芯片制造湿法清洗/刻蚀核心化学品，Stella垄断，涨幅>50%",
     "晶瑞电材","300655.SZ","深交所创业板","电子级化学品/光刻胶","CNY"),

    # ⑧MLCC电容
    ("⑧","MLCC多层陶瓷电容","电子工业大米，用量极大，今年涨幅>50%，全球供应缺口明显",
     "风华高科","000636.SZ","深交所","MLCC/片式电容","CNY"),
    ("⑧","MLCC多层陶瓷电容","电子工业大米，用量极大，今年涨幅>50%，全球供应缺口明显",
     "三环集团","300408.SZ","深交所创业板","MLCC/电子陶瓷","CNY"),
    ("⑧","MLCC多层陶瓷电容","电子工业大米，用量极大，今年涨幅>50%，全球供应缺口明显",
     "洁美科技","002859.SZ","深交所","MLCC载带/隔离纸","CNY"),

    # ⑨铜箔
    ("⑨","铜箔 Copper Foil","锂电负极/PCB覆铜板关键材料，今年价格翻倍，明年需求旺盛",
     "铜冠铜箔","301217.SZ","深交所","电解铜箔","CNY"),
    ("⑨","铜箔 Copper Foil","锂电负极/PCB覆铜板关键材料，今年价格翻倍，明年需求旺盛",
     "嘉元科技","688388.SH","上交所科创板","高端电解铜箔","CNY"),
    ("⑨","铜箔 Copper Foil","锂电负极/PCB覆铜板关键材料，今年价格翻倍，明年需求旺盛",
     "中一科技","003031.SZ","深交所","电解铜箔","CNY"),

    # ⑩电子布
    ("⑩","电子布 Electronic Fiberglass","覆铜板关键增强材料，高端电子布价格翻倍，缺货持续至2028",
     "中国巨石","600176.SH","上交所","玻纤/电子布","CNY"),
    ("⑩","电子布 Electronic Fiberglass","覆铜板关键增强材料，高端电子布价格翻倍，缺货持续至2028",
     "宏和科技","603256.SH","上交所","高端电子布","CNY"),
    ("⑩","电子布 Electronic Fiberglass","覆铜板关键增强材料，高端电子布价格翻倍，缺货持续至2028",
     "中材科技","002080.SZ","深交所","玻纤/复合材料","CNY"),

    # ⑪半导体靶材钽
    ("⑪","半导体靶材/钽 Sputtering Target","PVD溅射工艺不可或缺，今年涨幅80%，缺货至2027底",
     "金钼股份","601958.SH","上交所","钼靶材/贵金属","CNY"),
    ("⑪","半导体靶材/钽 Sputtering Target","PVD溅射工艺不可或缺，今年涨幅80%，缺货至2027底",
     "洛阳股份","000537.SZ","深交所","钼铜靶材","CNY"),
    ("⑪","半导体靶材/钽 Sputtering Target","PVD溅射工艺不可或缺，今年涨幅80%，缺货至2027底",
     "有研新材","600206.SH","上交所","半导体靶材/新材料","CNY"),

    # ⑫高纯氢气
    ("⑫","高纯氢气 Ultra-Pure H₂","半导体蚀刻/清洗/气相沉积核心气体，今年价格涨一倍",
     "凯美特气","002549.SZ","深交所","电子特种气体","CNY"),
    ("⑫","高纯氢气 Ultra-Pure H₂","半导体蚀刻/清洗/气相沉积核心气体，今年价格涨一倍",
     "华特气体","688268.SH","上交所科创板","电子特种气体","CNY"),
    ("⑫","高纯氢气 Ultra-Pure H₂","半导体蚀刻/清洗/气相沉积核心气体，今年价格涨一倍",
     "雅克科技","002409.SZ","深交所","半导体特种气体/前驱体","CNY"),
]

MAT_COLORS = {
    "①":"4472C4","②":"7030A0","③":"C00000","④":"833C00",
    "⑤":"0070C0","⑥":"375623","⑦":"ED7D31","⑧":"9B2335",
    "⑨":"C55A11","⑩":"538135","⑪":"203864","⑫":"4BACC6",
}


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — AI算力产业链 (2026年20大细分赛道)
# ═══════════════════════════════════════════════════════════════════════════════
# (赛道编号, 赛道名称, 排名, 公司名称, 股票代码, 市场, 细分定位, 币种)
AI_DATA = [
    # CPO
    ("01","CPO 共封装光学","Top1","中际旭创","300308.SZ","深交所","CPO/光模块","CNY"),
    ("01","CPO 共封装光学","Top2","新易盛","300502.SZ","深交所创业板","CPO/光模块","CNY"),
    ("01","CPO 共封装光学","Top3","天孚通信","300394.SZ","深交所创业板","CPO光互连组件","CNY"),
    ("01","CPO 共封装光学","Top4","华工科技","000988.SZ","深交所","激光/光通信器件","CNY"),

    # OCS
    ("02","OCS 光交换","Top1","腾景科技","688195.SH","上交所科创板","光开关/OCS器件","CNY"),
    ("02","OCS 光交换","Top2","福晶科技","002222.SZ","深交所","晶体光学/激光器件","CNY"),
    ("02","OCS 光交换","Top3","光库科技","300620.SZ","深交所创业板","OCS光学组件","CNY"),
    ("02","OCS 光交换","Top4","德科立","688205.SH","上交所科创板","光模块/相干光器件","CNY"),

    # 光芯片
    ("03","光芯片 Photonic Chip","Top1","源杰科技","688498.SH","上交所科创板","激光/光芯片","CNY"),
    ("03","光芯片 Photonic Chip","Top2","仕佳光子","688313.SH","上交所科创板","光芯片/PLC","CNY"),
    ("03","光芯片 Photonic Chip","Top3","光迅科技","002281.SZ","深交所","光芯片/光模块","CNY"),
    ("03","光芯片 Photonic Chip","Top4","长光华芯","688048.SH","上交所科创板","高功率激光芯片","CNY"),

    # PCB
    ("04","PCB 印刷电路板","Top1","胜宏科技","300476.SZ","深交所创业板","AI服务器PCB","CNY"),
    ("04","PCB 印刷电路板","Top2","东山精密","002384.SZ","深交所","高端PCB/FPC","CNY"),
    ("04","PCB 印刷电路板","Top3","深南电路","002916.SZ","深交所","AI服务器/封装基板","CNY"),
    ("04","PCB 印刷电路板","Top4","沪电股份","002463.SZ","深交所","高速高频PCB","CNY"),

    # AI服务器
    ("05","AI服务器","Top1","工业富联","601138.SH","上交所","AI服务器代工/组装","CNY"),
    ("05","AI服务器","Top2","浪潮信息","000977.SZ","深交所","AI服务器整机","CNY"),
    ("05","AI服务器","Top3","紫光股份","000938.SZ","深交所","AI服务器/网络设备","CNY"),
    ("05","AI服务器","Top4","中科曙光","603019.SH","上交所","AI服务器/超算","CNY"),

    # AI芯片
    ("06","AI芯片 AI Chip","Top1","海光信息","688041.SH","上交所科创板","GPU/DCU芯片","CNY"),
    ("06","AI芯片 AI Chip","Top2","寒武纪","688256.SH","上交所科创板","AI推理/训练芯片","CNY"),
    ("06","AI芯片 AI Chip","Top3","沐曦股份","非上市","—","GPU架构芯片","CNY"),
    ("06","AI芯片 AI Chip","Top4","摩尔线程","非上市","—","GPU芯片","CNY"),

    # 光纤光缆
    ("07","光纤光缆","Top1","长飞光纤","601869.SH","上交所","光纤预制棒/光缆","CNY"),
    ("07","光纤光缆","Top2","亨通光电","600487.SH","上交所","光纤光缆/海缆","CNY"),
    ("07","光纤光缆","Top3","中天科技","600522.SH","上交所","光纤光缆/海缆","CNY"),
    ("07","光纤光缆","Top4","烽火通信","600498.SH","上交所","光纤光缆/通信设备","CNY"),

    # 存储芯片
    ("08","存储芯片 Memory","Top1","德明利","001323.SZ","深交所","NAND Flash存储模组","CNY"),
    ("08","存储芯片 Memory","Top2","兆易创新","603986.SH","上交所","NOR Flash/MCU","CNY"),
    ("08","存储芯片 Memory","Top3","佰维存储","688525.SH","上交所科创板","NAND存储模组","CNY"),
    ("08","存储芯片 Memory","Top4","江波龙","301308.SZ","深交所","NAND/eMMC存储","CNY"),

    # 光模块设备
    ("09","光模块设备","Top1","罗博特科","688062.SH","上交所科创板","光模块自动化设备","CNY"),
    ("09","光模块设备","Top2","科瑞技术","002957.SZ","深交所","光模块检测设备","CNY"),
    ("09","光模块设备","Top3","博杰股份","300784.SZ","深交所创业板","光模块自动化/AOI","CNY"),
    ("09","光模块设备","Top4","德龙激光","688170.SH","上交所科创板","激光加工设备","CNY"),

    # 高速链接
    ("10","高速链接 High Speed Connect","Top1","立讯精密","002475.SZ","深交所","高速连接器/线缆","CNY"),
    ("10","高速链接 High Speed Connect","Top2","兆龙互连","603557.SH","上交所","高速铜缆/连接器","CNY"),
    ("10","高速链接 High Speed Connect","Top3","沃尔核材","002130.SZ","深交所","高速线缆/热缩材料","CNY"),
    ("10","高速链接 High Speed Connect","Top4","鼎通科技","301291.SZ","深交所","高速连接器","CNY"),

    # 钢箔/铜箔
    ("11","钢箔/铜箔 Copper Foil","Top1","诺德股份","600110.SH","上交所","电解铜箔","CNY"),
    ("11","钢箔/铜箔 Copper Foil","Top2","铜冠铜箔","301217.SZ","深交所","电解铜箔","CNY"),
    ("11","钢箔/铜箔 Copper Foil","Top3","嘉元科技","688388.SH","上交所科创板","高端电解铜箔","CNY"),
    ("11","钢箔/铜箔 Copper Foil","Top4","德福科技","非上市","—","电解铜箔","CNY"),

    # 树脂
    ("12","树脂 Resin","Top1","东材科技","601208.SH","上交所","特种树脂/覆铜板材料","CNY"),
    ("12","树脂 Resin","Top2","圣泉集团","605589.SH","上交所","酚醛树脂/电子材料","CNY"),
    ("12","树脂 Resin","Top3","美联新材","300586.SZ","深交所创业板","覆铜板树脂","CNY"),
    ("12","树脂 Resin","Top4","宏昌电子","603008.SH","上交所","环氧树脂/电子材料","CNY"),

    # 电子布
    ("13","电子布 Electronic Fiberglass","Top1","宏和科技","603256.SH","上交所","高端电子布","CNY"),
    ("13","电子布 Electronic Fiberglass","Top2","中国巨石","600176.SH","上交所","玻纤/电子布","CNY"),
    ("13","电子布 Electronic Fiberglass","Top3","中材科技","002080.SZ","深交所","玻纤/复合材料","CNY"),
    ("13","电子布 Electronic Fiberglass","Top4","国际复材","非上市","—","电子级玻纤布","CNY"),

    # 液冷
    ("14","液冷 Liquid Cooling","Top1","英维克","002837.SZ","深交所","液冷/机房热管理","CNY"),
    ("14","液冷 Liquid Cooling","Top2","高澜股份","300499.SZ","深交所创业板","液冷系统","CNY"),
    ("14","液冷 Liquid Cooling","Top3","中科曙光","603019.SH","上交所","液冷AI服务器/超算","CNY"),
    ("14","液冷 Liquid Cooling","Top4","中菱环境","非上市","—","液冷系统集成","CNY"),

    # 电源
    ("15","电源 Power Supply","Top1","中恒电气","002364.SZ","深交所","数据中心电源","CNY"),
    ("15","电源 Power Supply","Top2","圣阳股份","002580.SZ","深交所","UPS/储能电源","CNY"),
    ("15","电源 Power Supply","Top3","欧陆通","300870.SZ","深交所创业板","服务器电源","CNY"),
    ("15","电源 Power Supply","Top4","麦格米特","002851.SZ","深交所","工业电源/服务器电源","CNY"),

    # 燃气轮机
    ("16","燃气轮机 Gas Turbine","Top1","杰瑞股份","002353.SZ","深交所","燃气轮机/油气设备","CNY"),
    ("16","燃气轮机 Gas Turbine","Top2","联德股份","605060.SH","上交所","燃气轮机零部件","CNY"),
    ("16","燃气轮机 Gas Turbine","Top3","应流股份","603308.SH","上交所","航空/燃机叶片铸件","CNY"),
    ("16","燃气轮机 Gas Turbine","Top4","东方电气","600875.SH","上交所","燃气轮机/发电设备","CNY"),

    # 固态变压器
    ("17","固态变压器 SST","Top1","四方股份","601126.SH","上交所","固态变压器/电力电子","CNY"),
    ("17","固态变压器 SST","Top2","中国西电","601179.SH","上交所","高压电力设备","CNY"),
    ("17","固态变压器 SST","Top3","伊戈尔","002922.SZ","深交所","磁性元件/变压器","CNY"),
    ("17","固态变压器 SST","Top4","金盘科技","688676.SH","上交所科创板","干式变压器/电力设备","CNY"),

    # AIDC
    ("18","AIDC AI数据中心","Top1","润泽科技","300442.SZ","深交所","AIDC建设运营","CNY"),
    ("18","AIDC AI数据中心","Top2","网宿科技","300017.SZ","深交所创业板","CDN/AIDC","CNY"),
    ("18","AIDC AI数据中心","Top3","光环新网","300383.SZ","深交所创业板","IDC/AIDC","CNY"),
    ("18","AIDC AI数据中心","Top4","数据港","603881.SH","上交所","IDC/AIDC建设运营","CNY"),

    # 算电协同
    ("19","算电协同 Computing-Power Grid","Top1","豫能控股","001896.SZ","深交所","电力/算力协同","CNY"),
    ("19","算电协同 Computing-Power Grid","Top2","协鑫能科","002015.SZ","深交所","绿电/算力协同","CNY"),
    ("19","算电协同 Computing-Power Grid","Top3","南网数字","非上市","—","数字电网/算力","CNY"),
    ("19","算电协同 Computing-Power Grid","Top4","韶能股份","000601.SZ","深交所","绿电/数字能源","CNY"),

    # 算力租赁
    ("20","算力租赁 Computing Leasing","Top1","利通电子","603629.SH","上交所","算力租赁/GPU云","CNY"),
    ("20","算力租赁 Computing Leasing","Top2","协创数据","300456.SZ","深交所创业板","算力/数据中心租赁","CNY"),
    ("20","算力租赁 Computing Leasing","Top3","宏景科技","非上市","—","算力租赁","CNY"),
    ("20","算力租赁 Computing Leasing","Top4","优刻得","688158.SH","上交所科创板","算力云/GPU租赁","CNY"),
]

AI_TRACK_COLORS = {
    "01":"4472C4","02":"7030A0","03":"C00000","04":"375623",
    "05":"ED7D31","06":"833C00","07":"0070C0","08":"538135",
    "09":"9B2335","10":"203864","11":"C55A11","12":"4BACC6",
    "13":"7030A0","14":"00B0F0","15":"FFC000","16":"FF0000",
    "17":"92D050","18":"4472C4","19":"ED7D31","20":"375623",
}


# ═══════════════════════════════════════════════════════════════════════════════
# 构建 Sheet 1
# ═══════════════════════════════════════════════════════════════════════════════
def build_sheet1(wb):
    ws = wb.active
    ws.title = "①半导体制造设备"

    NCOLS = 5 + 4 + PRICE_COLS  # 阶段+步骤+类型+说明+公司+代码+市场+行业+币+参与度 + 价格
    TOTAL = 10 + PRICE_COLS     # 实际列数

    set_col_widths(ws, {
        "A": 22, "B": 6,  "C": 15, "D": 8,  "E": 32,
        "F": 28, "G": 12, "H": 14, "I": 20, "J": 6,
        "K": 6,  # 参与度
        "L": 11, "M": 11, "N": 12, "O": 12,
        "P": 13, "Q": 12, "R": 11, "S": 11, "T": 11,
    })

    header_row = write_title(
        ws,
        "半导体制造设备 Semiconductor Equipment — 按工序阶段排列",
        20,
        "类型：SEQ顺序  LOOP×N多层循环  PAR并行路径  NEW新兴探索   参与度：★主导/垄断(>50%)  ●核心参与  ○参与"
    )

    HEADERS = [
        "阶段 Stage", "步骤#", "步骤名称", "类型",
        "说明 Description",
        "标的名称", "标的代码", "市场", "主要/细分行业", "币种", "参与度",
    ] + PRICE_HEADERS

    write_header_row(ws, header_row, HEADERS)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:T{header_row}"

    stage_rows  = {}
    step_rows   = {}
    data_start  = header_row + 1

    for rn, rec in enumerate(EQUIP_DATA, start=data_start):
        stage, step_num, step_name, stype, desc, name, ticker, market, sub, cur, parti = rec
        ws.row_dimensions[rn].height = 17

        sc = STAGE_COLORS.get(stage, "808080")
        tc = TYPE_COLORS.get(stype, "FFFFFF")
        pc = PARTI_COLORS.get(parti, "FFFFFF")

        vals = [stage, step_num, step_name, stype, desc, name, ticker, market, sub, cur, parti] + [""]*PRICE_COLS
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=rn, column=ci, value=v)
            cell.border = BORDER
            if ci == 1:
                cell.fill, cell.font, cell.alignment = hf(sc), tf(bold=True, color="FFFFFF", size=9), C
            elif ci == 4:
                cell.fill = hf(tc)
                cell.font = tf(bold=True, color="000000" if stype == "NEW" else "FFFFFF", size=9)
                cell.alignment = C
            elif ci == 11:
                cell.fill = hf(pc)
                cell.font = tf(bold=True, color="FFFFFF" if parti in ("★","●") else "404040", size=10)
                cell.alignment = C
            elif ci in (2, 3):
                cell.fill = hf("F2F2F2")
                cell.font = tf(bold=(ci == 3), size=9)
                cell.alignment = C
            elif ci == 5:
                cell.font = tf(size=8, color="404040")
                cell.alignment = L
            elif ci == 6:
                cell.font = tf(bold=True, size=9)
                cell.alignment = L
            elif ci == 7:
                cell.font = Font(bold=True, color="1F3864", size=9, name="Courier New")
                cell.alignment = C
            elif ci == 10:
                cell.font = tf(size=9, color="595959")
                cell.alignment = C
            else:
                cell.font = tf(size=9)
                cell.alignment = C if ci >= 10 else L

        stage_rows.setdefault(stage, []).append(rn)
        step_key = (stage, step_num, step_name)
        step_rows.setdefault(step_key, []).append(rn)

    # 合并阶段列
    for stage, rows in stage_rows.items():
        if len(rows) > 1:
            ws.merge_cells(f"A{rows[0]}:A{rows[-1]}")
            ws[f"A{rows[0]}"].alignment = C

    # 合并步骤号/名/类型/说明
    for sk, rows in step_rows.items():
        if len(rows) > 1:
            for col in (2, 3, 4, 5):
                lc = get_column_letter(col)
                ws.merge_cells(f"{lc}{rows[0]}:{lc}{rows[-1]}")
                ws[f"{lc}{rows[0]}"].alignment = C


# ═══════════════════════════════════════════════════════════════════════════════
# 构建 Sheet 2
# ═══════════════════════════════════════════════════════════════════════════════
def build_sheet2(wb):
    ws = wb.create_sheet("②半导体关键材料")

    set_col_widths(ws, {
        "A": 8,  "B": 22, "C": 38,
        "D": 22, "E": 13, "F": 14, "G": 20, "H": 6,
        "I": 11, "J": 11, "K": 12, "L": 12,
        "M": 13, "N": 12, "O": 11, "P": 11, "Q": 11,
    })

    TOTAL = 8 + PRICE_COLS
    header_row = write_title(
        ws,
        "半导体12大卡脖子关键材料 — A股核心标的",
        TOTAL,
        "数据来源：万联摩尔AI-2026最新产业数据库    币种：CNY=人民币A股"
    )

    HEADERS = ["编号","材料类别","重要性/短缺说明","标的名称","标的代码","市场","主要/细分行业","币种"] + PRICE_HEADERS
    write_header_row(ws, header_row, HEADERS)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(TOTAL)}{header_row}"

    cat_rows = {}
    data_start = header_row + 1

    for rn, rec in enumerate(MATERIAL_DATA, start=data_start):
        num, cat, desc, name, ticker, market, sub, cur = rec
        ws.row_dimensions[rn].height = 17
        cc = MAT_COLORS.get(num, "808080")

        vals = [num, cat, desc, name, ticker, market, sub, cur] + [""]*PRICE_COLS
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=rn, column=ci, value=v)
            cell.border = BORDER
            if ci == 1:
                cell.fill = hf(cc)
                cell.font = tf(bold=True, color="FFFFFF", size=11)
                cell.alignment = C
            elif ci == 2:
                cell.fill = hf(cc + "40" if len(cc) == 6 else "EEF2FF")
                light = "D6E4F7"
                cell.fill = hf(light)
                cell.font = tf(bold=True, size=9, color=cc)
                cell.alignment = L
            elif ci == 3:
                cell.font = tf(size=8, color="404040", italic=True)
                cell.alignment = L
            elif ci == 4:
                cell.font = tf(bold=True, size=9)
                cell.alignment = L
            elif ci == 5:
                cell.font = Font(bold=True, color="1F3864", size=9, name="Courier New")
                cell.alignment = C
            elif ci == 8:
                cell.font = tf(size=9, color="595959")
                cell.alignment = C
            else:
                cell.font = tf(size=9)
                cell.alignment = C if ci >= 8 else L

        cat_rows.setdefault(num, []).append(rn)

    # 合并编号列 + 类别列 + 说明列
    for num, rows in cat_rows.items():
        if len(rows) > 1:
            for col in (1, 2, 3):
                lc = get_column_letter(col)
                ws.merge_cells(f"{lc}{rows[0]}:{lc}{rows[-1]}")
                ws[f"{lc}{rows[0]}"].alignment = C if col == 1 else L


# ═══════════════════════════════════════════════════════════════════════════════
# 构建 Sheet 3
# ═══════════════════════════════════════════════════════════════════════════════
def build_sheet3(wb):
    ws = wb.create_sheet("③AI算力产业链")

    set_col_widths(ws, {
        "A": 6,  "B": 22, "C": 6,
        "D": 22, "E": 13, "F": 14, "G": 20, "H": 6,
        "I": 11, "J": 11, "K": 12, "L": 12,
        "M": 13, "N": 12, "O": 11, "P": 11, "Q": 11,
    })

    TOTAL = 8 + PRICE_COLS
    header_row = write_title(
        ws,
        "AI算力产业链 2026年核心赛道20大细分领域",
        TOTAL,
        "数据来源：万联摩尔AI-2026最新产业数据库    标注：非上市=暂未上市公司"
    )

    HEADERS = ["赛道#","赛道名称","排名","标的名称","标的代码","市场","主要/细分行业","币种"] + PRICE_HEADERS
    write_header_row(ws, header_row, HEADERS)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(TOTAL)}{header_row}"

    track_rows = {}
    data_start = header_row + 1

    for rn, rec in enumerate(AI_DATA, start=data_start):
        num, track, rank, name, ticker, market, sub, cur = rec
        ws.row_dimensions[rn].height = 17
        cc = AI_TRACK_COLORS.get(num, "808080")

        # rank标签颜色
        rank_colors = {"Top1":"C00000","Top2":"ED7D31","Top3":"4472C4","Top4":"595959"}
        rc = rank_colors.get(rank, "808080")

        vals = [num, track, rank, name, ticker, market, sub, cur] + [""]*PRICE_COLS
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=rn, column=ci, value=v)
            cell.border = BORDER
            if ci == 1:
                cell.fill = hf(cc)
                cell.font = tf(bold=True, color="FFFFFF", size=10)
                cell.alignment = C
            elif ci == 2:
                cell.fill = hf("D6E4F7")
                cell.font = tf(bold=True, size=9, color=cc)
                cell.alignment = L
            elif ci == 3:
                cell.fill = hf(rc)
                cell.font = tf(bold=True, color="FFFFFF", size=9)
                cell.alignment = C
            elif ci == 4:
                cell.font = tf(bold=True, size=9)
                cell.alignment = L
            elif ci == 5:
                cell.font = Font(bold=True, color="1F3864", size=9, name="Courier New")
                cell.alignment = C
            elif ci == 8:
                cell.font = tf(size=9, color="595959")
                cell.alignment = C
            elif ticker == "非上市" and ci in (4, 5, 6):
                cell.fill = hf("F2F2F2")
                cell.font = tf(size=9, color="808080", italic=True)
                cell.alignment = C if ci >= 5 else L
            else:
                cell.font = tf(size=9)
                cell.alignment = C if ci >= 8 else L

        track_rows.setdefault(num, []).append(rn)

    # 合并赛道编号 + 赛道名称
    for num, rows in track_rows.items():
        if len(rows) > 1:
            for col in (1, 2):
                lc = get_column_letter(col)
                ws.merge_cells(f"{lc}{rows[0]}:{lc}{rows[-1]}")
                ws[f"{lc}{rows[0]}"].alignment = C if col == 1 else L


# ═══════════════════════════════════════════════════════════════════════════════
# 主入口
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    fname = "半导体全产业链标的.xlsx"
    wb = Workbook()
    build_sheet1(wb)
    build_sheet2(wb)
    build_sheet3(wb)
    wb.save(fname)

    # 统计
    s1 = len(EQUIP_DATA)
    s2 = len(MATERIAL_DATA)
    s3 = len(AI_DATA)
    print(f"[OK] 已生成: {fname}")
    print(f"  Sheet1 半导体设备: {s1} 行")
    print(f"  Sheet2 关键材料:   {s2} 行")
    print(f"  Sheet3 AI算力:     {s3} 行")
    print(f"  合计:              {s1+s2+s3} 行")

if __name__ == "__main__":
    main()
